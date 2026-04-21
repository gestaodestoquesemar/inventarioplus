"""
excel_generator.py — Gera relatório Excel formatado com múltiplas abas.
"""
import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint


# ── Paleta ─────────────────────────────────────────────────────────────────────
VERDE_ESCURO  = "0F2A1E"
VERDE_MED     = "1D9E75"
VERDE_CLARO   = "E1F5EE"
VERMELHO      = "E24B4A"
VERMELHO_CLARO= "FCEBEB"
AMBAR         = "BA7517"
AMBAR_CLARO   = "FAEEDA"
AZUL          = "378ADD"
CINZA_CLARO   = "F1EFE8"
CINZA_MEDIO   = "888780"
BRANCO        = "FFFFFF"


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _fonte(bold=False, color="444441", size=9, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)


def _borda():
    lado = Side(style="thin", color="D0D0D0")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _centralizado(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _esquerda(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _ajustar_colunas(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 4))


def _cabecalho_aba(ws, titulo, subtitulo=""):
    """Insere cabeçalho verde escuro na aba."""
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = titulo
    c.font = Font(bold=True, color=BRANCO, size=14, name="Calibri")
    c.fill = _fill(VERDE_ESCURO)
    c.alignment = _centralizado()
    ws.row_dimensions[1].height = 30

    if subtitulo:
        ws.merge_cells("A2:H2")
        c2 = ws["A2"]
        c2.value = subtitulo
        c2.font = Font(color=CINZA_MEDIO, size=9, name="Calibri")
        c2.fill = _fill(CINZA_CLARO)
        c2.alignment = _centralizado()
        ws.row_dimensions[2].height = 18
        return 3
    return 2


def _linha_cabecalho_tabela(ws, row, colunas):
    for col_idx, texto in enumerate(colunas, start=1):
        c = ws.cell(row=row, column=col_idx, value=texto)
        c.font = _fonte(bold=True, color=BRANCO, size=9)
        c.fill = _fill(VERDE_MED)
        c.alignment = _centralizado(wrap=True)
        c.border = _borda()
    ws.row_dimensions[row].height = 22


def _linha_dados(ws, row, valores, bg=BRANCO, negrito=False):
    for col_idx, val in enumerate(valores, start=1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font = _fonte(bold=negrito)
        c.fill = _fill(bg)
        c.border = _borda()
        c.alignment = _esquerda(wrap=True)
    ws.row_dimensions[row].height = 18


# ═══════════════════════════════════════════════════════════
# ABA 1 — RESUMO EXECUTIVO
# ═══════════════════════════════════════════════════════════
def _aba_resumo(wb, resumo, diagnosticos, observacoes):
    ws = wb.create_sheet("Resumo Executivo")
    ws.sheet_view.showGridLines = False

    linha = _cabecalho_aba(ws, "RESUMO EXECUTIVO — INVENTÁRIO",
        f"Gerado em {datetime.today().strftime('%d/%m/%Y %H:%M')}")

    linha += 1
    kpis = [
        ("Perda Total (R$)", f"R$ {abs(resumo['perda_total']):,.2f}", VERMELHO),
        ("% Perda / Estoque", f"{resumo['pct_perda']:.2f}%", AMBAR if resumo["pct_perda"] > 2 else VERDE_MED),
        ("Ganhos — Ajustes Positivos", f"R$ {resumo['ganho_total']:,.2f}", VERDE_MED),
        ("Total de Itens Auditados", f"{resumo['total_itens']:,}", AZUL),
        ("Itens com Perda", str(resumo["n_perdas"]), VERMELHO),
        ("Itens com Ganho", str(resumo["n_ganhos"]), VERDE_MED),
        ("Itens com Ajuste", str(resumo["n_ajustes"]), AZUL),
        ("Estoque Total Avaliado (R$)", f"R$ {resumo['estoque_total']:,.2f}", CINZA_MEDIO),
    ]

    ws.merge_cells(f"A{linha}:H{linha}")
    ws[f"A{linha}"].value = "INDICADORES PRINCIPAIS"
    ws[f"A{linha}"].font = _fonte(bold=True, size=10, color=VERDE_ESCURO)
    ws[f"A{linha}"].fill = _fill(VERDE_CLARO)
    ws[f"A{linha}"].alignment = _esquerda()
    linha += 1

    _linha_cabecalho_tabela(ws, linha, ["Indicador", "Valor"])
    ws.merge_cells(f"C{linha}:H{linha}")
    linha += 1

    for i, (nome, val, cor) in enumerate(kpis):
        bg = CINZA_CLARO if i % 2 == 0 else BRANCO
        ws.cell(row=linha, column=1, value=nome).font = _fonte(bold=True)
        ws.cell(row=linha, column=1).fill = _fill(bg)
        ws.cell(row=linha, column=1).border = _borda()
        ws.cell(row=linha, column=1).alignment = _esquerda()

        c_val = ws.cell(row=linha, column=2, value=val)
        c_val.font = Font(bold=True, color=cor, size=10, name="Calibri")
        c_val.fill = _fill(bg)
        c_val.border = _borda()
        c_val.alignment = _centralizado()

        ws.merge_cells(f"C{linha}:H{linha}")
        ws.cell(row=linha, column=3).fill = _fill(bg)
        linha += 1

    linha += 1
    ws.merge_cells(f"A{linha}:H{linha}")
    ws[f"A{linha}"].value = "DIAGNÓSTICO AUTOMÁTICO"
    ws[f"A{linha}"].font = _fonte(bold=True, size=10, color=VERDE_ESCURO)
    ws[f"A{linha}"].fill = _fill(VERDE_CLARO)
    ws[f"A{linha}"].alignment = _esquerda()
    linha += 1

    for d in diagnosticos.get("alertas_criticos", []) + diagnosticos.get("alertas_moderados", []):
        ws.merge_cells(f"A{linha}:H{linha}")
        ws[f"A{linha}"].value = f"⚑ {d['titulo']}"
        ws[f"A{linha}"].font = _fonte(bold=True, color=VERMELHO)
        ws[f"A{linha}"].fill = _fill(VERMELHO_CLARO)
        ws[f"A{linha}"].alignment = _esquerda(wrap=True)
        ws.row_dimensions[linha].height = 20
        linha += 1

        ws.merge_cells(f"A{linha}:H{linha}")
        ws[f"A{linha}"].value = d["descricao"]
        ws[f"A{linha}"].font = _fonte(color="555555")
        ws[f"A{linha}"].fill = _fill(BRANCO)
        ws[f"A{linha}"].alignment = _esquerda(wrap=True)
        ws.row_dimensions[linha].height = 30
        linha += 1

    linha += 1
    ws.merge_cells(f"A{linha}:H{linha}")
    ws[f"A{linha}"].value = "RECOMENDAÇÕES"
    ws[f"A{linha}"].font = _fonte(bold=True, size=10, color=VERDE_ESCURO)
    ws[f"A{linha}"].fill = _fill(VERDE_CLARO)
    ws[f"A{linha}"].alignment = _esquerda()
    linha += 1

    cor_prior = {"ALTA": VERMELHO, "MÉDIA": AMBAR, "BAIXA": VERDE_MED}
    for rec in diagnosticos.get("recomendacoes", []):
        ws.merge_cells(f"A{linha}:H{linha}")
        ws[f"A{linha}"].value = f"[{rec['prioridade']}] {rec['texto']}"
        ws[f"A{linha}"].font = _fonte(color=cor_prior.get(rec["prioridade"], CINZA_MEDIO))
        ws[f"A{linha}"].fill = _fill(BRANCO)
        ws[f"A{linha}"].alignment = _esquerda(wrap=True)
        ws.row_dimensions[linha].height = 20
        linha += 1

    if observacoes and observacoes.strip():
        linha += 1
        ws.merge_cells(f"A{linha}:H{linha}")
        ws[f"A{linha}"].value = "OBSERVAÇÕES DO ANALISTA"
        ws[f"A{linha}"].font = _fonte(bold=True, size=10, color=VERDE_ESCURO)
        ws[f"A{linha}"].fill = _fill(VERDE_CLARO)
        ws[f"A{linha}"].alignment = _esquerda()
        linha += 1

        ws.merge_cells(f"A{linha}:H{linha}")
        ws[f"A{linha}"].value = observacoes
        ws[f"A{linha}"].font = _fonte(color="555555")
        ws[f"A{linha}"].alignment = _esquerda(wrap=True)
        ws.row_dimensions[linha].height = max(30, len(observacoes) // 5)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22


# ═══════════════════════════════════════════════════════════
# ABA 2 — TOP PERDAS
# ═══════════════════════════════════════════════════════════
def _aba_top_perdas(wb, top_perdas):
    ws = wb.create_sheet("Top Perdas")
    ws.sheet_view.showGridLines = False
    linha = _cabecalho_aba(ws, "TOP 10 — MAIORES PERDAS FINANCEIRAS",
        "Produtos com maior impacto financeiro negativo no inventário")

    colunas = ["#", "Código", "Produto", "Departamento", "Qtd Ajustada", "Valor Perda (R$)", "Causa Estimada"]
    _linha_cabecalho_tabela(ws, linha, colunas)
    linha += 1

    for i, row in top_perdas.iterrows():
        bg = VERMELHO_CLARO if i % 2 == 0 else BRANCO
        vals = [
            i + 1, str(row["codigo"]), str(row["descricao"]),
            str(row["departamento"]), int(row["quantidade"]),
            round(abs(row["valor"]), 2), str(row["causa_estimada"])
        ]
        _linha_dados(ws, linha, vals, bg=bg)
        ws.cell(row=linha, column=6).font = Font(bold=True, color=VERMELHO, size=9, name="Calibri")
        ws.cell(row=linha, column=6).number_format = 'R$ #,##0.00'
        ws.cell(row=linha, column=6).alignment = _centralizado()
        ws.cell(row=linha, column=1).alignment = _centralizado()
        ws.cell(row=linha, column=5).alignment = _centralizado()
        linha += 1

    _ajustar_colunas(ws)
    ws.column_dimensions["C"].width = 40


# ═══════════════════════════════════════════════════════════
# ABA 3 — POR DEPARTAMENTO
# ═══════════════════════════════════════════════════════════
def _aba_departamentos(wb, por_depto):
    ws = wb.create_sheet("Por Departamento")
    ws.sheet_view.showGridLines = False
    linha = _cabecalho_aba(ws, "RANKING DE PERDAS POR DEPARTAMENTO",
        "Ordenado do mais crítico para o menos crítico")

    colunas = ["Departamento", "Perda Total (R$)", "% do Total", "Classificação"]
    _linha_cabecalho_tabela(ws, linha, colunas)
    linha += 1

    for idx, (_, row) in enumerate(por_depto.iterrows()):
        bg = VERMELHO_CLARO if idx == 0 else (AMBAR_CLARO if idx <= 2 else BRANCO)
        classif = "🔴 Crítico" if idx == 0 else ("🟡 Atenção" if idx <= 2 else "🟢 Estável")
        vals = [
            str(row["departamento"]),
            round(abs(row["perda"]), 2),
            row["pct"],
            classif
        ]
        _linha_dados(ws, linha, vals, bg=bg)
        ws.cell(row=linha, column=2).number_format = 'R$ #,##0.00'
        ws.cell(row=linha, column=2).font = Font(bold=True, color=VERMELHO, size=9, name="Calibri")
        ws.cell(row=linha, column=2).alignment = _centralizado()
        ws.cell(row=linha, column=3).number_format = '0.0"%"'
        ws.cell(row=linha, column=3).alignment = _centralizado()
        ws.cell(row=linha, column=4).alignment = _centralizado()
        linha += 1

    _ajustar_colunas(ws)


# ═══════════════════════════════════════════════════════════
# ABA 4 — DADOS COMPLETOS
# ═══════════════════════════════════════════════════════════
def _aba_dados_completos(wb, df):
    ws = wb.create_sheet("Dados Completos")
    ws.sheet_view.showGridLines = False
    linha = _cabecalho_aba(ws, "DADOS COMPLETOS DO INVENTÁRIO",
        "Todos os registros processados com classificação automática")

    colunas = ["Código", "Descrição", "Departamento", "Data", "Qtd Ajustada", "Valor (R$)", "Tipo", "Causa Estimada"]
    _linha_cabecalho_tabela(ws, linha, colunas)
    linha += 1

    for i, row in df.iterrows():
        bg = VERMELHO_CLARO if row["tipo"] == "perda" else (VERDE_CLARO if row["tipo"] == "ganho" else BRANCO)
        data_str = row["data"].strftime("%d/%m/%Y") if hasattr(row["data"], "strftime") and pd.notna(row["data"]) else ""
        vals = [
            str(row["codigo"]), str(row["descricao"]),
            str(row["departamento"]), data_str,
            int(row["quantidade"]), round(row["valor"], 2),
            str(row["tipo"]).upper(), str(row["causa_estimada"])
        ]
        _linha_dados(ws, linha, vals, bg=bg)
        cor_val = VERMELHO if row["tipo"] == "perda" else (VERDE_MED if row["tipo"] == "ganho" else CINZA_MEDIO)
        ws.cell(row=linha, column=6).font = Font(bold=True, color=cor_val, size=9, name="Calibri")
        ws.cell(row=linha, column=6).number_format = 'R$ #,##0.00'
        ws.cell(row=linha, column=6).alignment = _centralizado()
        ws.cell(row=linha, column=5).alignment = _centralizado()
        linha += 1

    ws.auto_filter.ref = f"A3:{get_column_letter(len(colunas))}{linha - 1}"
    _ajustar_colunas(ws)
    ws.column_dimensions["B"].width = 40


# ═══════════════════════════════════════════════════════════
# FUNÇÃO PRINCIPAL
# ═══════════════════════════════════════════════════════════
def gerar_excel_relatorio(resultado, observacoes=""):
    resumo       = resultado["resumo"]
    top_perdas   = resultado["top_perdas"]
    top_ganhos   = resultado["top_ganhos"]
    por_depto    = resultado["por_depto"]
    diagnosticos = resultado["diagnosticos"]
    df           = resultado["df"]

    wb = Workbook()
    wb.remove(wb.active)

    _aba_resumo(wb, resumo, diagnosticos, observacoes)
    _aba_top_perdas(wb, top_perdas)
    _aba_departamentos(wb, por_depto)
    _aba_dados_completos(wb, df)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
